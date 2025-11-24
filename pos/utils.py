from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
from django.utils import timezone
from decimal import Decimal


def generate_receipt_pdf(order):
    """Generate thermal receipt PDF (58mm width)"""
    buffer = BytesIO()

    # 58mm width = 58mm, height flexible
    width = 58 * mm
    height = 200 * mm  # Will adjust based on content

    # Create PDF
    p = canvas.Canvas(buffer, pagesize=(width, height))

    # Starting Y position (from top)
    y = height - 10 * mm

    # Store name and header
    p.setFont("Helvetica-Bold", 10)
    p.drawCentredString(width / 2, y, "MINI POS")
    y -= 4 * mm

    p.setFont("Helvetica", 7)
    p.drawCentredString(width / 2, y, "Jl. Contoh No. 123")
    y -= 3 * mm
    p.drawCentredString(width / 2, y, "Telp: 08123456789")
    y -= 5 * mm

    # Separator line
    p.line(5 * mm, y, width - 5 * mm, y)
    y -= 4 * mm

    # Order info
    p.setFont("Helvetica", 7)
    p.drawString(5 * mm, y, f"No: #{order.id}")
    y -= 3 * mm
    p.drawString(5 * mm, y, f"Tanggal: {order.created_at.strftime('%d/%m/%Y %H:%M')}")
    y -= 3 * mm
    p.drawString(5 * mm, y, f"Pelanggan: {order.customer.name}")
    y -= 4 * mm

    # Separator
    p.line(5 * mm, y, width - 5 * mm, y)
    y -= 4 * mm

    # Items header
    p.setFont("Helvetica-Bold", 7)
    p.drawString(5 * mm, y, "Item")
    p.drawRightString(width - 5 * mm, y, "Subtotal")
    y -= 3 * mm

    p.setFont("Helvetica", 6)

    # Items
    for item in order.items.all():
        # Product name (might wrap if too long)
        product_name = item.product.name
        if len(product_name) > 25:
            product_name = product_name[:22] + "..."

        p.drawString(5 * mm, y, product_name)
        y -= 3 * mm

        # Quantity, price, discount
        qty_price = f"{item.quantity} x Rp {item.price:,.0f}"
        p.drawString(7 * mm, y, qty_price)
        y -= 3 * mm

        if item.discount_percent > 0:
            discount_text = (
                f"Diskon {item.discount_percent}% (-Rp {item.discount_amount():,.0f})"
            )
            p.drawString(7 * mm, y, discount_text)
            y -= 3 * mm

        # Subtotal
        subtotal_text = f"Rp {item.subtotal():,.0f}"
        p.drawRightString(width - 5 * mm, y, subtotal_text)
        y -= 4 * mm

    # Separator
    p.line(5 * mm, y, width - 5 * mm, y)
    y -= 4 * mm

    # Total
    p.setFont("Helvetica-Bold", 8)
    p.drawString(5 * mm, y, "TOTAL")
    p.drawRightString(width - 5 * mm, y, f"Rp {order.total_price:,.0f}")
    y -= 5 * mm

    # Separator
    p.line(5 * mm, y, width - 5 * mm, y)
    y -= 5 * mm

    # Footer
    p.setFont("Helvetica", 6)
    p.drawCentredString(width / 2, y, "Terima kasih atas kunjungan Anda")
    y -= 3 * mm
    p.drawCentredString(width / 2, y, "Barang yang sudah dibeli")
    y -= 3 * mm
    p.drawCentredString(width / 2, y, "tidak dapat ditukar/dikembalikan")

    # Finalize PDF
    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer


def generate_report_pdf(start_date, end_date, orders, summary):
    """Generate report PDF in A4 format"""
    buffer = BytesIO()

    # Create PDF with A4 page
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Starting position
    y = height - 50

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, y, "LAPORAN PENJUALAN")
    y -= 20

    # Period
    p.setFont("Helvetica", 11)
    period_text = (
        f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    )
    p.drawCentredString(width / 2, y, period_text)
    y -= 30

    # Summary box
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Ringkasan:")
    y -= 20

    p.setFont("Helvetica", 10)
    p.drawString(70, y, f"Total Transaksi: {summary['total_orders']}")
    y -= 15
    p.drawString(70, y, f"Total Penjualan: Rp {summary['total_sales']:,.0f}")
    y -= 15
    p.drawString(70, y, f"Rata-rata per Transaksi: Rp {summary['avg_sales']:,.0f}")
    y -= 30

    # Table header
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "No")
    p.drawString(100, y, "Tanggal")
    p.drawString(200, y, "Pelanggan")
    p.drawString(400, y, "Total")
    y -= 5
    p.line(50, y, width - 50, y)
    y -= 15

    # Table data
    p.setFont("Helvetica", 9)
    for idx, order in enumerate(orders, 1):
        if y < 100:  # New page if needed
            p.showPage()
            y = height - 50
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y, "No")
            p.drawString(100, y, "Tanggal")
            p.drawString(200, y, "Pelanggan")
            p.drawString(400, y, "Total")
            y -= 5
            p.line(50, y, width - 50, y)
            y -= 15
            p.setFont("Helvetica", 9)

        p.drawString(50, y, str(idx))
        p.drawString(100, y, order.created_at.strftime("%d/%m/%Y %H:%M"))

        # Truncate customer name if too long
        customer_name = order.customer.name
        if len(customer_name) > 30:
            customer_name = customer_name[:27] + "..."
        p.drawString(200, y, customer_name)

        p.drawRightString(width - 50, y, f"Rp {order.total_price:,.0f}")
        y -= 15

    # Footer
    y -= 20
    p.line(50, y, width - 50, y)
    y -= 20
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, y, "TOTAL")
    p.drawRightString(width - 50, y, f"Rp {summary['total_sales']:,.0f}")

    # Finalize
    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer


def generate_report_excel(start_date, end_date, orders, summary):
    """Generate report Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Title
    ws["A1"] = "LAPORAN PENJUALAN"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:D1")

    # Period
    ws["A2"] = (
        f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")

    # Summary
    ws["A4"] = "Ringkasan"
    ws["A4"].font = Font(bold=True, size=12)

    ws["A5"] = "Total Transaksi:"
    ws["B5"] = summary["total_orders"]

    ws["A6"] = "Total Penjualan:"
    ws["B6"] = f"Rp {summary['total_sales']:,.0f}"

    ws["A7"] = "Rata-rata per Transaksi:"
    ws["B7"] = f"Rp {summary['avg_sales']:,.0f}"

    # Table header
    row = 9
    headers = ["No", "Tanggal", "Pelanggan", "Total"]
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Table data
    row += 1
    for idx, order in enumerate(orders, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=order.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=3, value=order.customer.name)
        ws.cell(row=row, column=4, value=f"Rp {order.total_price:,.0f}")
        row += 1

    # Total row
    row += 1
    total_cell = ws.cell(row=row, column=1, value="TOTAL")
    total_cell.font = Font(bold=True)
    ws.merge_cells(f"A{row}:C{row}")

    total_value_cell = ws.cell(
        row=row, column=4, value=f"Rp {summary['total_sales']:,.0f}"
    )
    total_value_cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
