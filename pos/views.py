from decimal import Decimal
import json
from datetime import timedelta, datetime
from django.utils import timezone
from django.db.models import Sum, Count, F
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import (
    Product,
    Customer,
    Order,
    OrderItem,
    Category,
    Supplier,
    PurchaseOrder,
    PurchaseOrderItem,
)
from .forms import (
    ProductForm,
    CustomerForm,
    CategoryForm,
    SupplierForm,
)
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from django.http import HttpResponseForbidden
from functools import wraps
from django.contrib import messages
from .utils import generate_receipt_pdf, generate_report_pdf, generate_report_excel


def login_view(request):
    """Login page for authentication."""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get("next", "dashboard")
            return redirect(next_url)
        else:
            messages.error(request, "Username atau password salah.")

    return render(request, "pos/login.html")


def logout_view(request):
    """Logout user and redirect to login."""
    logout(request)
    messages.success(request, "Anda telah logout.")
    return redirect("login")


def clear_cache_page(request):
    """Page to clear PWA cache and service worker."""
    return render(request, "clear_cache.html")


@login_required
def dashboard(request):
    """Show a small dashboard with counts."""
    products_count = Product.objects.count()
    customers_count = Customer.objects.count()
    orders_count = Order.objects.count()
    # total revenue (sum of all orders)
    total_rev = Order.objects.aggregate(total=Sum("total_price"))["total"] or Decimal(
        "0"
    )

    # Low stock warning (products with stock < 5)
    low_stock_products = Product.objects.filter(stock__lt=5).order_by("stock")
    low_stock_count = low_stock_products.count()

    # last 7 days sales (by order.created_at date)
    today = timezone.localdate()
    labels = []
    data = []
    labels_display = []
    counts = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        labels.append(day.strftime("%Y-%m-%d"))
        # human-friendly label, e.g. '22 Nov'
        labels_display.append(day.strftime("%d %b"))
        day_total = Order.objects.filter(created_at__date=day).aggregate(
            total=Sum("total_price")
        )["total"] or Decimal("0")
        day_count = (
            Order.objects.filter(created_at__date=day).aggregate(count=Count("id"))[
                "count"
            ]
            or 0
        )
        # convert Decimal to float for JSON/Chart.js
        data.append(float(day_total))
        counts.append(day_count)

    return render(
        request,
        "pos/dashboard.html",
        {
            "products_count": products_count,
            "customers_count": customers_count,
            "orders_count": orders_count,
            "total_revenue": total_rev,
            "low_stock_products": low_stock_products[:5],  # Show top 5
            "low_stock_count": low_stock_count,
            "sales_labels_json": json.dumps(labels),
            "sales_labels_display_json": json.dumps(labels_display),
            "sales_counts_json": json.dumps(counts),
            "sales_data_json": json.dumps(data),
        },
    )


@login_required
def low_stock_products(request):
    """Show products with low stock (< 5)."""
    products = Product.objects.filter(stock__lt=5).order_by("stock", "name")

    return render(
        request,
        "pos/low_stock_products.html",
        {
            "products": products,
            "low_stock_count": products.count(),
        },
    )


@login_required
def category_list(request):
    """List all categories."""
    categories = Category.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        categories = categories.filter(name__icontains=query)
    return render(
        request, "pos/category_list.html", {"categories": categories, "query": query}
    )


def category_create(request):
    """Create a new category."""
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("category_list")
    else:
        form = CategoryForm()
    return render(request, "pos/category_form.html", {"form": form})


@login_required
def category_update(request, pk):
    """Update existing category."""
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect("category_list")
    else:
        form = CategoryForm(instance=category)
    return render(
        request, "pos/category_form.html", {"form": form, "category": category}
    )


def category_delete(request, pk):
    """Delete a category."""
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        return redirect("category_list")
    return render(request, "pos/category_confirm_delete.html", {"category": category})


@login_required
def product_list(request):
    """List all products."""
    products = Product.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        products = products.filter(name__icontains=query)
    products = products.order_by("-created_at")
    return render(
        request, "pos/product_list.html", {"products": products, "query": query}
    )


@login_required
def product_create(request):
    """Create a new product."""
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(request, "pos/product_form.html", {"form": form})


@login_required
def product_update(request, pk):
    """Update existing product."""
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect("product_list")
    else:
        form = ProductForm(instance=product)
    return render(request, "pos/product_form.html", {"form": form, "product": product})


@login_required
def product_delete(request, pk):
    """Delete a product."""
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.delete()
        return redirect("product_list")
    return render(request, "pos/product_confirm_delete.html", {"product": product})


@login_required
def customer_list(request):
    """List customers."""
    customers = Customer.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        customers = customers.filter(name__icontains=query) | customers.filter(
            phone__icontains=query
        )
    customers = customers.order_by("-created_at")
    return render(
        request, "pos/customer_list.html", {"customers": customers, "query": query}
    )


@login_required
def customer_create(request):
    """Create a new customer."""
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("customer_list")
    else:
        form = CustomerForm()
    return render(request, "pos/customer_form.html", {"form": form})


@login_required
def customer_update(request, pk):
    """Update a customer."""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect("customer_list")
    else:
        form = CustomerForm(instance=customer)
    return render(
        request, "pos/customer_form.html", {"form": form, "customer": customer}
    )


@login_required
def customer_delete(request, pk):
    """Delete a customer."""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == "POST":
        customer.delete()
        return redirect("customer_list")
    return render(request, "pos/customer_confirm_delete.html", {"customer": customer})


@login_required
def supplier_list(request):
    """List all suppliers."""
    suppliers = Supplier.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        suppliers = suppliers.filter(name__icontains=query)
    return render(
        request, "pos/supplier_list.html", {"suppliers": suppliers, "query": query}
    )


@login_required
def supplier_create(request):
    """Create a new supplier."""
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("supplier_list")
    else:
        form = SupplierForm()
    return render(request, "pos/supplier_form.html", {"form": form})


@login_required
def supplier_update(request, pk):
    """Update existing supplier."""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect("supplier_list")
    else:
        form = SupplierForm(instance=supplier)
    return render(
        request, "pos/supplier_form.html", {"form": form, "supplier": supplier}
    )


@login_required
def supplier_delete(request, pk):
    """Delete a supplier."""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        supplier.delete()
        return redirect("supplier_list")
    return render(request, "pos/supplier_confirm_delete.html", {"supplier": supplier})


@login_required
def purchase_order_list(request):
    """List purchase orders."""
    purchase_orders = PurchaseOrder.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        purchase_orders = purchase_orders.filter(
            order_number__icontains=query
        ) | purchase_orders.filter(supplier__name__icontains=query)
    return render(
        request,
        "pos/purchase_order_list.html",
        {"purchase_orders": purchase_orders, "query": query},
    )


@login_required
def purchase_order_create(request):
    """Create purchase order with items."""
    suppliers = Supplier.objects.all()
    products = Product.objects.all()

    if request.method == "POST":
        supplier_id = request.POST.get("supplier")
        order_number = request.POST.get("order_number")
        notes = request.POST.get("notes", "")
        product_ids = request.POST.getlist("product")
        qtys = request.POST.getlist("quantity")
        prices = request.POST.getlist("unit_price")

        if not supplier_id or not order_number:
            return HttpResponseBadRequest("Missing supplier or order number")

        supplier = get_object_or_404(Supplier, pk=supplier_id)

        items = []
        for pid, q, p in zip(product_ids, qtys, prices):
            try:
                prod = Product.objects.get(pk=int(pid))
                qty = int(q)
                unit_price = Decimal(p)
            except Exception:
                continue
            if qty <= 0 or unit_price <= 0:
                continue
            items.append((prod, qty, unit_price))

        if not items:
            return HttpResponseBadRequest("No valid items")

        with transaction.atomic():
            total = Decimal("0")
            for prod, qty, unit_price in items:
                total += qty * unit_price

            po = PurchaseOrder.objects.create(
                supplier=supplier,
                order_number=order_number,
                total_amount=total,
                notes=notes,
            )

            for prod, qty, unit_price in items:
                PurchaseOrderItem.objects.create(
                    purchase_order=po,
                    product=prod,
                    quantity=qty,
                    unit_price=unit_price,
                )

        return redirect("purchase_order_detail", pk=po.id)

    return render(
        request,
        "pos/purchase_order_create.html",
        {"suppliers": suppliers, "products": products},
    )


@login_required
def purchase_order_detail(request, pk):
    """Show purchase order detail."""
    po = get_object_or_404(PurchaseOrder, pk=pk)
    return render(request, "pos/purchase_order_detail.html", {"purchase_order": po})


@login_required
def purchase_order_receive(request, pk):
    """Receive purchase order and update stock."""
    po = get_object_or_404(PurchaseOrder, pk=pk)

    if po.status != "pending":
        messages.error(request, "Purchase order sudah diproses sebelumnya.")
        return redirect("purchase_order_detail", pk=po.id)

    if request.method == "POST":
        with transaction.atomic():
            # Update stock for each item
            for item in po.items.all():
                item.product.stock += item.quantity
                item.product.save(update_fields=["stock"])

            # Update PO status
            po.status = "received"
            po.received_date = timezone.now()
            po.save(update_fields=["status", "received_date"])

            messages.success(
                request,
                f"Purchase Order {po.order_number} berhasil diterima. Stok telah diperbarui.",
            )

        return redirect("purchase_order_detail", pk=po.id)

    return render(request, "pos/purchase_order_receive.html", {"purchase_order": po})


@login_required
def orders_list(request):
    """List orders."""
    orders = Order.objects.all()
    query = request.GET.get("q", "").strip()
    if query:
        orders = orders.filter(customer__name__icontains=query)
    orders = orders.order_by("-created_at")
    return render(request, "pos/orders_list.html", {"orders": orders, "query": query})


@login_required
def order_detail(request, pk):
    """Show order detail."""
    order = get_object_or_404(Order, pk=pk)
    return render(request, "pos/order_detail.html", {"order": order})


@login_required
def order_receipt(request, pk):
    """Generate and download receipt PDF."""
    order = get_object_or_404(Order, pk=pk)
    pdf_buffer = generate_receipt_pdf(order)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="receipt_{order.id}.pdf"'
    return response


@login_required
def order_create(request):
    """Create order from form POST with multiple items.

    POST fields: 'customer', multiple 'product', 'quantity', and 'discount'
    """
    customers = Customer.objects.all()
    products = Product.objects.all()

    if request.method == "POST":
        customer_id = request.POST.get("customer")
        product_ids = request.POST.getlist("product")
        qtys = request.POST.getlist("quantity")
        discounts = request.POST.getlist("discount")

        if not customer_id:
            return HttpResponseBadRequest("Missing customer")

        customer = get_object_or_404(Customer, pk=customer_id)

        items = []
        insufficient = []
        for pid, q, disc in zip(product_ids, qtys, discounts):
            try:
                prod = Product.objects.get(pk=int(pid))
                qty = int(q)
                discount = Decimal(disc) if disc else Decimal("0")
            except Exception:
                continue
            if qty <= 0:
                continue
            # Check stock availability; collect insufficient instead of 400
            if prod.stock < qty:
                insufficient.append((prod, qty))
                continue
            items.append((prod, qty, discount))

        if insufficient:
            # Show error messages for each insufficient product
            for prod, qty in insufficient:
                messages.error(
                    request,
                    f"Stok tidak cukup untuk produk '{prod.name}'. Diminta {qty}, tersedia {prod.stock}.",
                )
            return render(
                request,
                "pos/order_create.html",
                {"customers": customers, "products": products},
            )

        if not items:
            return HttpResponseBadRequest("No valid items")

        with transaction.atomic():
            total = Decimal("0")
            for prod, qty, discount in items:
                subtotal = prod.price * qty
                discount_amount = subtotal * (discount / 100)
                total += subtotal - discount_amount

            order = Order.objects.create(customer=customer, total_price=total)
            out_of_stock = []
            for prod, qty, discount in items:
                OrderItem.objects.create(
                    order=order,
                    product=prod,
                    quantity=qty,
                    price=prod.price,
                    discount_percent=discount,
                )
                # Decrease product stock and save
                prod.stock = prod.stock - qty
                prod.save(update_fields=["stock"])
                if prod.stock == 0:
                    out_of_stock.append(prod.name)

        if out_of_stock:
            # Show one warning listing all products that are now empty
            messages.warning(
                request,
                "Stok habis untuk: "
                + ", ".join(out_of_stock)
                + ". Mohon restok segera.",
            )

        return redirect("order_detail", pk=order.id)

    return render(
        request, "pos/order_create.html", {"customers": customers, "products": products}
    )


@login_required
def reports(request):
    """Simple transactions report with period filter."""
    # Get period filter (default: all)
    period = request.GET.get("period", "all")

    orders_qs = Order.objects.all()

    # Apply date filter based on period
    today = timezone.localdate()
    start_date = None
    end_date = today

    if period == "daily":
        start_date = today
        orders_qs = orders_qs.filter(created_at__date=today)
    elif period == "weekly":
        start_date = today - timedelta(days=7)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    elif period == "monthly":
        start_date = today - timedelta(days=30)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)

    orders_qs = orders_qs.order_by("-created_at")

    # total sales across filtered orders (Decimal)
    total_sales = orders_qs.aggregate(total=Sum("total_price"))["total"] or Decimal("0")

    # server-side pagination for reports (page size 15)
    page_size = 15
    paginator = Paginator(orders_qs, page_size)
    page = request.GET.get("page")
    try:
        orders_page = paginator.page(page)
    except PageNotAnInteger:
        orders_page = paginator.page(1)
    except EmptyPage:
        orders_page = paginator.page(paginator.num_pages)

    return render(
        request,
        "pos/reports.html",
        {
            "orders": orders_page,
            "total_sales": total_sales,
            "paginator": paginator,
            "page_obj": orders_page,
            "is_paginated": orders_page.has_other_pages(),
            "period": period,
            "start_date": start_date,
            "end_date": end_date,
        },
    )


@login_required
def report_export_pdf(request):
    """Export report to PDF"""
    period = request.GET.get("period", "all")

    orders_qs = Order.objects.all()

    # Apply date filter
    today = timezone.localdate()
    start_date = None
    end_date = today

    if period == "daily":
        start_date = today
        orders_qs = orders_qs.filter(created_at__date=today)
    elif period == "weekly":
        start_date = today - timedelta(days=7)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    elif period == "monthly":
        start_date = today - timedelta(days=30)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    else:
        # For 'all', use earliest order date as start
        first_order = orders_qs.order_by("created_at").first()
        start_date = first_order.created_at.date() if first_order else today

    orders_qs = orders_qs.order_by("-created_at")

    # Calculate summary
    total_sales = orders_qs.aggregate(total=Sum("total_price"))["total"] or Decimal("0")
    total_orders = orders_qs.count()
    avg_sales = total_sales / total_orders if total_orders > 0 else Decimal("0")

    summary = {
        "total_sales": total_sales,
        "total_orders": total_orders,
        "avg_sales": avg_sales,
    }

    # Generate PDF
    pdf_buffer = generate_report_pdf(start_date, end_date, orders_qs, summary)

    # Return as download
    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    filename = f'laporan_{period}_{today.strftime("%Y%m%d")}.pdf'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def report_export_excel(request):
    """Export report to Excel"""
    period = request.GET.get("period", "all")

    orders_qs = Order.objects.all()

    # Apply date filter
    today = timezone.localdate()
    start_date = None
    end_date = today

    if period == "daily":
        start_date = today
        orders_qs = orders_qs.filter(created_at__date=today)
    elif period == "weekly":
        start_date = today - timedelta(days=7)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    elif period == "monthly":
        start_date = today - timedelta(days=30)
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    else:
        # For 'all', use earliest order date as start
        first_order = orders_qs.order_by("created_at").first()
        start_date = first_order.created_at.date() if first_order else today

    orders_qs = orders_qs.order_by("-created_at")

    # Calculate summary
    total_sales = orders_qs.aggregate(total=Sum("total_price"))["total"] or Decimal("0")
    total_orders = orders_qs.count()
    avg_sales = total_sales / total_orders if total_orders > 0 else Decimal("0")

    summary = {
        "total_sales": total_sales,
        "total_orders": total_orders,
        "avg_sales": avg_sales,
    }

    # Generate Excel
    excel_buffer = generate_report_excel(start_date, end_date, orders_qs, summary)

    # Return as download
    response = HttpResponse(
        excel_buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f'laporan_{period}_{today.strftime("%Y%m%d")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def analytics(request):
    """Advanced analytics with charts - optimized version."""
    today = timezone.localdate()
    start_date = today - timedelta(days=7)  # Reduced from 30 to 7 days

    # Top 5 categories only
    category_sales = (
        OrderItem.objects.filter(order__created_at__date__gte=start_date)
        .values(category_name=F("product__category__name"))
        .annotate(
            total=Sum(F("quantity") * F("price") * (1 - F("discount_percent") / 100))
        )
        .order_by("-total")[:5]  # Limit to top 5
    )

    category_labels = []
    category_data = []
    for item in category_sales:
        if item["category_name"]:
            category_labels.append(item["category_name"])
            category_data.append(float(item["total"] or 0))

    # Top 5 products only
    top_products = (
        OrderItem.objects.filter(order__created_at__date__gte=start_date)
        .values(product_name=F("product__name"))
        .annotate(
            total_qty=Sum("quantity"),
            total_sales=Sum(
                F("quantity") * F("price") * (1 - F("discount_percent") / 100)
            ),
        )
        .order_by("-total_sales")[:5]  # Reduced from 10 to 5
    )

    product_labels = []
    product_qty_data = []
    product_sales_data = []
    for item in top_products:
        product_labels.append(item["product_name"])
        product_qty_data.append(int(item["total_qty"] or 0))
        product_sales_data.append(float(item["total_sales"] or 0))

    # Simplified hourly sales - only last 7 days
    hourly_sales = (
        Order.objects.filter(created_at__date__gte=start_date)
        .annotate(hour=F("created_at__hour"))
        .values("hour")
        .annotate(total=Sum("total_price"), count=Count("id"))
        .order_by("hour")
    )

    hour_labels = [f"{h:02d}:00" for h in range(24)]
    hour_data = [0] * 24
    hour_counts = [0] * 24

    for item in hourly_sales:
        hour = item["hour"]
        hour_data[hour] = float(item["total"] or 0)
        hour_counts[hour] = int(item["count"] or 0)

    return render(
        request,
        "pos/analytics.html",
        {
            "category_labels_json": json.dumps(category_labels),
            "category_data_json": json.dumps(category_data),
            "product_labels_json": json.dumps(product_labels),
            "product_qty_data_json": json.dumps(product_qty_data),
            "product_sales_data_json": json.dumps(product_sales_data),
            "hour_labels_json": json.dumps(hour_labels),
            "hour_data_json": json.dumps(hour_data),
            "hour_counts_json": json.dumps(hour_counts),
            "start_date": start_date,
            "end_date": today,
        },
    )


@login_required
def backups_list(request):
    """List available SQLite backups in backup/ directory."""
    backup_dir = settings.BASE_DIR / "backup"
    backup_dir.mkdir(exist_ok=True)
    files = []
    for f in sorted(backup_dir.glob("db_backup_*.sqlite3"), reverse=True):
        stat = f.stat()
        files.append(
            {
                "name": f.name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime),
            }
        )
    return render(request, "pos/backups.html", {"files": files})


@login_required
def backup_download(request, filename):
    """Download chosen backup file."""
    import re
    from django.http import FileResponse

    if not re.match(r"^db_backup_\d{8}\.sqlite3$", filename):
        return HttpResponseBadRequest("Invalid backup filename")
    backup_dir = settings.BASE_DIR / "backup"
    path = backup_dir / filename
    if not path.exists():
        return HttpResponseBadRequest("Backup not found")
    return FileResponse(open(path, "rb"), as_attachment=True, filename=filename)


@login_required
def product_import(request):
    """Import products from Excel file."""
    if request.method == "POST":
        if "file" not in request.FILES:
            messages.error(request, "Tidak ada file yang diupload.")
            return redirect("product_import")

        excel_file = request.FILES["file"]
        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Format file harus Excel (.xlsx atau .xls).")
            return redirect("product_import")

        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_file)
            ws = wb.active

            # Expected columns: Name, Category, Price, Stock, Description
            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            created_count = 0
            skipped = []

            for idx, row in enumerate(rows, start=2):
                if not row or not row[0]:  # Skip empty rows
                    continue
                try:
                    name = str(row[0]).strip()
                    category_name = str(row[1]).strip() if row[1] else ""
                    price = Decimal(str(row[2]))
                    stock = int(row[3]) if row[3] else 0
                    description = str(row[4]).strip() if row[4] and len(row) > 4 else ""

                    # Find or skip category
                    category = None
                    if category_name:
                        category, _ = Category.objects.get_or_create(name=category_name)

                    # Create product
                    Product.objects.create(
                        name=name,
                        category=category,
                        price=price,
                        stock=stock,
                        description=description,
                    )
                    created_count += 1
                except Exception as e:
                    skipped.append(f"Baris {idx}: {str(e)}")

            if created_count > 0:
                messages.success(
                    request, f"Berhasil mengimport {created_count} produk."
                )
            if skipped:
                messages.warning(
                    request,
                    f"Dilewati {len(skipped)} baris. Detail: {'; '.join(skipped[:5])}",
                )

        except Exception as e:
            messages.error(request, f"Error saat membaca file: {str(e)}")

        return redirect("product_list")

    return render(request, "pos/product_import.html")


@login_required
def product_import_template(request):
    """Download sample Excel template for product import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row
    headers = ["Name", "Category", "Price", "Stock", "Description"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Sample data
    ws.append(["Produk Contoh 1", "Elektronik", 150000, 10, "Deskripsi contoh"])
    ws.append(["Produk Contoh 2", "Makanan", 25000, 50, "Snack enak"])

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="template_import_produk.xlsx"'
    )
    return response


def require_api_key(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        # Check X-API-KEY header (also accept HTTP_X_API_KEY for older servers)
        key = request.headers.get("X-API-KEY") or request.META.get("HTTP_X_API_KEY")
        if not key or key != getattr(settings, "API_KEY", None):
            return HttpResponseForbidden("Invalid or missing API key")
        return view_func(request, *args, **kwargs)

    return _wrapped


@require_api_key
def api_products(request):
    """Return JSON list of products."""
    data = list(Product.objects.values("id", "name", "price", "stock", "description"))
    return JsonResponse({"products": data})


@require_api_key
@require_http_methods(["POST"])
def api_create_order(request):
    """Create an order from JSON POST.

    Expected JSON: {"customer": id, "items": [{"product": id, "quantity": n}, ...]}
    """
    import json

    try:
        payload = json.loads(request.body)
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    customer_id = payload.get("customer")
    items = payload.get("items", [])
    if not customer_id or not items:
        return HttpResponseBadRequest("Missing fields")

    customer = get_object_or_404(Customer, pk=customer_id)

    # Validate stock and prepare items
    items_data = []
    for it in items:
        pid = it.get("product")
        qty = int(it.get("quantity", 0))
        if not pid or qty <= 0:
            continue
        prod = get_object_or_404(Product, pk=pid)
        if prod.stock < qty:
            return HttpResponseBadRequest(
                f"Insufficient stock for product: {prod.name}"
            )
        items_data.append((prod, qty))

    with transaction.atomic():
        total = Decimal("0")
        for prod, qty in items_data:
            total += prod.price * qty

        order = Order.objects.create(customer=customer, total_price=total)
        for prod, qty in items_data:
            OrderItem.objects.create(
                order=order, product=prod, quantity=qty, price=prod.price
            )
            # decrease stock
            prod.stock = prod.stock - qty
            prod.save(update_fields=["stock"])

    return JsonResponse({"status": "ok", "order_id": order.id})
